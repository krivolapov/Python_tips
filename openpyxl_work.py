from openpyxl import load_workbook

wb_form = load_workbook(filename = 'Test_document.xlsx')
wb_val = load_workbook(filename = 'Test_document.xlsx', data_only = True)

sheet_form = wb_form['Data']
sheet_val = wb_val['Data']

E5_form = sheet_form['E5'].Value
E5_val = sheet_val['E5'].Value

print(E5_form)
print(E5_val)
